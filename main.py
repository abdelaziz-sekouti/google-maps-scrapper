import asyncio
import random
import re
import os
import argparse
import pandas as pd
from typing import List, Dict, Optional
from playwright.async_api import async_playwright, Page, BrowserContext
from rich.console import Console
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.logging import RichHandler
import logging

# Setup Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
    datefmt="[%X]",
    handlers=[RichHandler(rich_tracebacks=True), logging.FileHandler("log.txt")]
)
logger = logging.getLogger("scraper")
console = Console()

class GoogleMapsScraper:
    def __init__(self, target_leads: int = 100, headless: bool = True, indicatif: str = "212", city: str = "", message_template: str = ""):
        self.target_leads = target_leads
        self.headless = headless
        self.indicatif = indicatif.strip("+")
        self.city = city
        self.message_template = message_template
        self.leads = []
        self.processed_names = set()

    def normalize_phone(self, phone: str) -> str:
        """Normalizes phone numbers based on country indicatif"""
        if not phone:
            return ""
        
        # Remove all non-digit characters except +
        clean = re.sub(r'[^\d+]', '', phone)
        
        # Handle country code
        if clean.startswith(f'+{self.indicatif}'):
            clean = clean[len(self.indicatif)+1:]
        elif clean.startswith(self.indicatif):
            clean = clean[len(self.indicatif):]
        
        # Remove leading 0 if present (common in Europe/Africa)
        if clean.startswith('0'):
            clean = clean[1:]
            
        return clean

    def is_valid_phone(self, phone: str, filters: Dict) -> bool:
        """Checks if phone matches both whitelist and blacklist filters"""
        if not phone:
            return False
        
        exclude_prefixes = filters.get('exclude_prefixes', [])
        include_prefixes = filters.get('include_prefixes', [])

        # 1. Blacklist Check (Exclude)
        for prefix in exclude_prefixes:
            p_clean = prefix.strip("+").strip("0")
            if phone.startswith(p_clean) or phone.startswith(prefix):
                return False

        # 2. Whitelist Check (Include Only) - If whitelist is provided, phone MUST match one
        if include_prefixes:
            match_found = False
            for prefix in include_prefixes:
                p_clean = prefix.strip("+").strip("0")
                if phone.startswith(p_clean) or phone.startswith(prefix):
                    match_found = True
                    break
            if not match_found:
                return False

        return True


    def generate_wa_link(self, phone: str, name: str = "") -> str:
        """Generates WhatsApp link with current indicatif and pre-filled message"""
        if not phone:
            return ""
        
        base_url = f"https://wa.me/{self.indicatif}{phone}"
        
        if self.message_template:
            import urllib.parse
            msg = self.message_template.replace("[Name]", name).replace("[City]", self.city)
            encoded_msg = urllib.parse.quote(msg)
            return f"{base_url}?text={encoded_msg}"
        
        return base_url



    async def scroll_feed(self, page: Page):
        """Scrolls the left-side feed to load more listings"""
        # Try multiple common selectors for the scrollable feed
        selectors = ["div[role='feed']", "div.m67qEc", "div.section-layout", "div.ScrollableContainer"]
        feed = None
        for selector in selectors:
            try:
                el = page.locator(selector).first
                if await el.is_visible():
                    feed = el
                    break
            except:
                continue

        if feed:
            try:
                # Use evaluate to scroll
                await feed.evaluate("el => el.scrollTo(0, el.scrollHeight)")
                await asyncio.sleep(random.uniform(2.0, 4.0))
            except Exception as e:
                logger.debug(f"Scroll evaluation failed: {e}")
        else:
            # Fallback: simple page down on the whole page
            await page.keyboard.press("PageDown")
            await asyncio.sleep(2.0)

    async def extract_socials(self, page: Page) -> Dict[str, str]:
        """Attempts to find social links in the detail panel"""
        socials = {
            "instagram": "",
            "twitter": "",
            "facebook": "",
            "messenger": ""
        }
        
        links = await page.query_selector_all("a[href]")
        for link in links:
            href = await link.get_attribute("href")
            if not href: continue
            
            if "instagram.com" in href: socials["instagram"] = href
            elif "twitter.com" in href or "x.com" in href: socials["twitter"] = href
            elif "facebook.com" in href: socials["facebook"] = href
            elif "messenger.com" in href: socials["messenger"] = href
            
        return socials

    async def scrape(self, url: str, filters: Dict, progress_callback=None):
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=self.headless)
            context = await browser.new_context(
                viewport={'width': 1280, 'height': 800},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            )
            page = await context.new_page()
            
            # Sanitize URL: Take the last valid URL if multiple are joined
            urls = re.findall(r'https?://[^\s<>"]+', url)
            final_url = urls[-1] if urls else url
            
            logger.info(f"Navigating to: {final_url}")
            if progress_callback: progress_callback(f"Navigating to sanitized URL...")
            
            # Use 'load' instead of 'networkidle' to avoid timeouts on heavy maps
            try:
                await page.goto(final_url, wait_until="load", timeout=60000)
                await asyncio.sleep(5) # Manual buffer
            except Exception as e:
                logger.warning(f"Initial load timed out, attempting to proceed: {e}")
            
            # Handle cookie consent if it appears
            try:
                consent_btn = page.locator("button:has-text('Accept all'), button:has-text('Accepter'), button:has-text('Agree')").first
                if await consent_btn.is_visible():
                    await consent_btn.click()
                    await asyncio.sleep(2)
            except:
                pass

            if not progress_callback:
                with Progress(
                    SpinnerColumn(),
                    TextColumn("[progress.description]{task.description}"),
                    BarColumn(),
                    TaskProgressColumn(),
                    console=console
                ) as progress:
                    task = progress.add_task(f"Extracting leads...", total=self.target_leads)
                    await self._scrape_loop(page, filters, task, progress, None)
            else:
                await self._scrape_loop(page, filters, None, None, progress_callback)
                            
            await browser.close()
            return self.leads

    async def _scrape_loop(self, page, filters, task, progress, progress_callback):
        consecutive_duplicates = 0
        
        while len(self.leads) < self.target_leads:
            # 1. Get all visible listings
            listing_selector = "a.hfpxzc"
            listings = page.locator(listing_selector)
            count = await listings.count()
            
            if count == 0:
                logger.warning("No listings found. Trying to scroll...")
                if progress_callback: progress_callback("Scrolling to find listings...")
                await self.scroll_feed(page)
                # Check if we are at the end
                if await page.locator("text='You've reached the end of the list'").is_visible():
                    logger.info("Reached the end of Google Maps results.")
                    break
                continue

            # 2. Iterate through listings by index to avoid stale elements
            found_new_in_this_batch = False
            for i in range(count):
                if len(self.leads) >= self.target_leads:
                    break
                
                try:
                    listing = listings.nth(i)
                    # Scroll element into view if needed
                    await listing.scroll_into_view_if_needed()
                    
                    name_raw = await listing.get_attribute("aria-label")
                    if not name_raw: continue
                    
                    if name_raw in self.processed_names:
                        continue
                    
                    self.processed_names.add(name_raw)
                    found_new_in_this_batch = True
                    consecutive_duplicates = 0
                    
                    # Click and wait
                    await listing.click()
                    await asyncio.sleep(random.uniform(2, 3))
                    
                    # Wait for detail panel
                    try:
                        await page.wait_for_selector("div[role='main']", timeout=5000)
                    except:
                        pass
                    
                    # Extract Data
                    name = name_raw

                    
                    # --- Multi-Strategy Detail Extraction ---
                    
                    # 1. Address
                    address = ""
                    addr_selectors = ["button[data-item-id='address']", "div.Io6YTe.fontBodyMedium", "button[data-tooltip='Copy address']"]
                    for sel in addr_selectors:
                        try:
                            el = page.locator(sel).first
                            if await el.is_visible():
                                address = await el.inner_text()
                                if address: break
                        except: continue
                    
                    # 2. Website
                    website = ""
                    web_selectors = ["a[data-item-id='authority']", "a[data-tooltip='Open website']", "a[aria-label*='website']"]
                    for sel in web_selectors:
                        try:
                            el = page.locator(sel).first
                            if await el.is_visible():
                                website = await el.get_attribute("href")
                                if website: break
                        except: continue
                    
                    # 3. Phone (The trickiest one)
                    phone_raw = ""
                    phone_selectors = [
                        "button[data-item-id*='phone']", 
                        "button[aria-label^='Phone:']", 
                        "button[data-tooltip*='phone']",
                        "div.fontBodyMedium:has-text('+212')",
                        "div.fontBodyMedium:has-text('06')",
                        "div.fontBodyMedium:has-text('07')"
                    ]
                    for sel in phone_selectors:
                        try:
                            # Use locator for better filtering
                            el = page.locator(sel).first
                            if await el.is_visible():
                                phone_raw = await el.inner_text()
                                if phone_raw: break
                        except: continue
                    
                    # Fallback for phone: Search all labels in the panel
                    if not phone_raw:
                        try:
                            all_text = await page.locator("div[role='main']").inner_text()
                            # Use current indicatif for search, and match various spacing/dash formats
                            # Pattern matches +[indicatif] or 0 followed by 8-10 digits with common separators
                            phones = re.findall(rf'(\+{self.indicatif}[\s.-]?\d{{1,2}}[\s.-]?\d{{2}}[\s.-]?\d{{2}}[\s.-]?\d{{2}}|0\d[\s.-]?\d{{2}}[\s.-]?\d{{2}}[\s.-]?\d{{2}})', all_text)
                            if phones: phone_raw = phones[0]
                        except: pass
                    
                    # --- End of Detail Extraction ---
                    
                    normalized_phone = self.normalize_phone(phone_raw)
                    wa_link = self.generate_wa_link(normalized_phone, name)
                    status = "Has Website" if website else "No Website"
                    maps_link = page.url
                    
                    socials = await self.extract_socials(page)
                    
                    message = ""
                    if self.message_template:
                        message = self.message_template.replace("[Name]", name).replace("[City]", self.city)

                    lead_data = {
                        "Name": name,
                        "City": self.city,
                        "Indicatif": self.indicatif,
                        "Phone": normalized_phone,
                        "Status": status,
                        "WhatsApp Link": wa_link,
                        "Message": message,
                        "Website": website,
                        "Address": address,
                        **socials,
                        "Google Maps Link": maps_link
                    }


                    
                    if filters.get('include_website_only') and not website:
                        continue
                    
                    # Add lead if valid
                    if normalized_phone and self.is_valid_phone(normalized_phone, filters):
                        self.leads.append(lead_data)
                        if progress: progress.update(task, advance=1)
                        if progress_callback: progress_callback(f"Found: {name}", len(self.leads))
                        logger.info(f"Generated Lead: {name} | Phone: {normalized_phone}")
                    else:
                        logger.info(f"Skipped (No valid phone): {name}")

                    # --- Return to List (Special handle for small screens/layouts) ---
                    try:
                        back_btn = page.locator("button[aria-label='Back'], button[aria-label='Retour']").first
                        if await back_btn.is_visible():
                            await back_btn.click()
                            await asyncio.sleep(1)
                    except:
                        pass
                    
                except Exception as e:
                    logger.warning(f"Error extracting listing index {i}: {e}")
                    continue

            # 3. If no new leads were found in this batch of visible items, scroll!
            if not found_new_in_this_batch:
                consecutive_duplicates += 1
                logger.info(f"Scrolling to load more results (batch complete)...")
                if progress_callback: progress_callback("Scrolling for deeper results...")
                await self.scroll_feed(page)
                
                # If we've scrolled many times and found nothing new, we might be at the end
                if consecutive_duplicates > 10 or await page.locator("text='Reached the end'").count() > 0:
                    logger.info("No new leads found after multiple scrolls. Stopping.")
                    break
        
        return self.leads


    def export_data(self, filename_base: str = "leads"):
        if not self.leads:
            logger.warning("No data to export.")
            return

        final_df = pd.DataFrame(self.leads)
        
        # ------------------------
        # FRESH EXPORT MODE (No Append)
        csv_file = f"{filename_base}.csv"
        xlsx_file = f"{filename_base}.xlsx"
        # ------------------------

        # Reorder columns for better visibility
        cols = ["Name", "Phone", "City", "Message", "WhatsApp Link", "Status", "Indicatif", "Website", "Address"]
        other_cols = [c for c in final_df.columns if c not in cols]
        final_df = final_df[cols + other_cols]

        # Save CSV
        final_df.to_csv(csv_file, index=False)
        logger.info(f"Successfully saved to {csv_file} (Total: {len(final_df)} leads)")
        
        # Save Excel with HYPERLINK FORMULA
        import openpyxl
        from openpyxl.styles import Font, Color, Alignment, PatternFill
        
        with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='Leads')
            
            workbook = writer.book
            worksheet = writer.sheets['Leads']
            
            # Styling headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Apply Hyperlink formula to the WhatsApp Link column (Column 5)
            blue_font = Font(color="0000FF", underline="single")
            wa_col_idx = 5
            for row in range(2, len(final_df) + 2):
                link_val = worksheet.cell(row=row, column=wa_col_idx).value
                if link_val:
                    formula = f'=HYPERLINK("{link_val}", "🚀 SEND WHATSAPP")'
                    cell = worksheet.cell(row=row, column=wa_col_idx)
                    cell.value = formula
                    cell.font = blue_font

            # Auto-adjust columns width
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                worksheet.column_dimensions[column].width = min(max_length + 2, 50)
            
            worksheet.freeze_panes = 'A2'
            
        logger.info(f"Successfully saved to {xlsx_file}")
        self.print_preview(final_df)


    def print_preview(self, df: pd.DataFrame):
        table = Table(title="Leads Preview (Top 10)")
        
        # Add columns
        show_cols = ["Name", "Phone", "Status", "Website"]
        for col in show_cols:
            table.add_column(col, style="cyan")
            
        # Add rows (top 10)
        for _, row in df.head(10).iterrows():
            table.add_row(
                str(row["Name"]),
                str(row["Phone"]),
                str(row["Status"]),
                str(row["Website"])[:30] + "..." if len(str(row["Website"])) > 30 else str(row["Website"])
            )
        
        console.print(table)


async def main():
    parser = argparse.ArgumentParser(description="Google Maps Business Scraper")
    parser.add_argument("url", help="Google Maps search URL")
    parser.add_argument("--count", type=int, default=100, help="Number of leads to scrape (default: 100)")
    parser.add_argument("--website", action="store_true", help="Include only businesses with a website")
    parser.add_argument("--exclude-prefixes", type=str, help="Comma-separated prefixes to exclude (e.g. +2125,05)")
    parser.add_argument("--visible", action="store_false", dest="headless", help="Run browser in visible mode")
    parser.set_defaults(headless=True)
    
    args = parser.parse_args()
    
    filters = {
        "include_website_only": args.website,
        "exclude_prefixes": args.exclude_prefixes.split(",") if args.exclude_prefixes else []
    }
    
    scraper = GoogleMapsScraper(target_leads=args.count, headless=args.headless)
    
    try:
        leads = await scraper.scrape(args.url, filters)
        scraper.export_data()
        console.print(f"\n[bold green]✅ Scraping completed! Collected {len(leads)} leads.[/bold green]")
    except KeyboardInterrupt:
        console.print("\n[bold yellow]Stopping and saving current progress...[/bold yellow]")
        scraper.export_data()
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        scraper.export_data()

if __name__ == "__main__":
    asyncio.run(main())
